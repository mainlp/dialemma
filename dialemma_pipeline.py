import sys
import json
import gzip
import os.path
import multiprocessing
import spacy
from functools import partial
from collections import defaultdict
from collections import Counter
from Levenshtein import distance
from random import sample, seed, shuffle

# from nltk import RegexpTokenizer
from tqdm import tqdm
from nltk.tokenize import sent_tokenize

from openpyxl import Workbook
# from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText


seed(1)

is_debug = False
print(f"building annotation files (debug={is_debug})")

# https://spacy.io/models/de
nlp = spacy.load("de_core_news_lg")

# levensthein number of neighest neigbors
ld_knn = 10

# k most frequent german terms
de_topk = 500 if is_debug else 20_000

# number of random occurrences for each bar term
n_rand_bar_sentences = 3

# number of files
n_chunks = 1 if is_debug else 20
chunk_size = de_topk // n_chunks

# used for multiprocessing
cpus = multiprocessing.cpu_count()

# path where downloaded wikipedia cirrus files are stored
input_path = ""

# path where annotation files (.xlsx) are stored
output_path = input_path 


def get_raw_corpus(fPath):
    print("loading raw corpus")
    id_article = []
    for i, jsonl in tqdm(enumerate(gzip.open(fPath, "rt"))):
        jsonl = json.loads(jsonl)

        if i >= 5_000 and is_debug:
            break

        if i % 2 == 0:
            if "index" in jsonl:
                aid = jsonl["index"]["_id"] if "index" in jsonl else None
        else:
            content_model = jsonl["content_model"]
            if aid is not None and content_model == "wikitext":
                article = jsonl["text"]
                id_article.append((aid, article))

    return id_article


def extract(aid_article, do_lemmatize):
    aid, article = aid_article
    asid_to_sent = {}
    word_to_asids = defaultdict(list)
    word_to_pos_distr = dict()

    for sid, sent in enumerate(sent_tokenize(article, language="german")):
        asid = str(aid) + "-" + str(sid)
        asid_to_sent[asid] = sent
        for token in nlp(sent):
            tok = token.lemma_ if do_lemmatize else token.text
            if tok.isalpha(): # v3: get rid of length constraint "and len(tok) > 2:"
                word_to_asids[tok].append(asid)
                if tok not in word_to_pos_distr:
                    word_to_pos_distr[tok] = defaultdict(int)
                word_to_pos_distr[tok][token.pos_] += 1

    return [asid_to_sent, word_to_asids, word_to_pos_distr]


def extract_words_sents(id_article, page_language):
    do_lemmatize = True if page_language == "de" else False
    print(f"do_lemmatize={do_lemmatize}")

    fn = partial(extract, do_lemmatize=do_lemmatize)

    all_asid_to_sent = {}
    all_word_to_asids = defaultdict(list)
    all_word_to_pos_distr = defaultdict(Counter)

    with multiprocessing.Pool(processes=cpus) as pool:
        for asid_to_sent, word_to_asids, word_to_pos_distr in tqdm(pool.imap_unordered(fn, id_article), total=len(id_article)):
            all_asid_to_sent.update(asid_to_sent)

            for word, asids in word_to_asids.items():
                all_word_to_asids[word].extend(asids)

            for word, pos_distr in word_to_pos_distr.items():
                for pos, freq in pos_distr.items():
                    all_word_to_pos_distr[word][pos] += freq

    return all_asid_to_sent, all_word_to_asids, all_word_to_pos_distr


def filter_de(de_word2sids):
    tmp = [(w, len(sids)) for w, sids in de_word2sids.items()]
    tmp.sort(key=lambda x: x[1], reverse=True)
    return tmp[:de_topk]


def compute_ld_knn(de_term, bar_vocab):
    knn = sorted([(bar_term, distance(de_term, bar_term)) for bar_term in bar_vocab], key=lambda elem: elem[1])[:ld_knn]
    return de_term, knn


def main():
    lang = "bar"
    id_article = get_raw_corpus(os.path.join(input_path, "barwiki-20250310-cirrussearch-content.json.gz"))
    bar_id2sent, bar_word2sids, bar_pos = extract_words_sents(id_article, lang)

    lang = "de"
    id_article = get_raw_corpus(os.path.join(input_path, "dewiki-20250310-cirrussearch-content.json.gz"))
    de_id2sent, de_word2sids, de_pos = extract_words_sents(id_article, lang)

    # extract full German vocabulary
    de_full_vocab = {t for t, _ in de_word2sids.items()}

    # filter infrequent german vocabulary
    de_word2freq = filter_de(de_word2sids)
    assert len(dict(de_word2freq)) == len(de_word2freq)
    de_word2freq = dict(de_word2freq)

    de_filtered_vocab = set(de_word2freq.keys())

    # remove de terms from bar vocab
    bar_word2sids = {w: sids for w, sids in bar_word2sids.items() if w not in de_full_vocab}

    # keep n random bar sentences and record bar word frequencies
    print("subsambling random sentences")
    bar_word2_freq = {}
    for w in bar_word2sids.keys():
        freq = len(bar_word2sids[w])
        bar_word2_freq[w] = freq
        bar_word2sids[w] = sample(bar_word2sids[w], min(n_rand_bar_sentences, freq))

    # extract ld-knns for each de term
    print("extracting knn's")
    de2bar_knn = {}
    knn_fn = partial(compute_ld_knn, bar_vocab=set(bar_word2_freq.keys()))
    with multiprocessing.Pool(processes=cpus) as pool:
        for de_term, knns in tqdm(pool.imap_unordered(knn_fn, de_filtered_vocab), total=len(de_filtered_vocab)):
            de2bar_knn[de_term] = knns


    print("composing annotation records")
    vocab_list = list(de_filtered_vocab)
    chunks = [vocab_list[i:i + chunk_size] for i in range(0, len(vocab_list), chunk_size)]

    _id = 0

    for chunk_i, chunk_vocab in enumerate(chunks, start=1):
        print(f"writing file for chunk: {chunk_i} out of {n_chunks}")

        wb = Workbook()
        ws1 = wb.worksheets[0]
        de_term_counter = 1
        row = 1

        ws1[f"A{row}"] = "id"

        ws1[f"B{row}"] = "de"
        ws1[f"C{row}"] = "freq_de"
        ws1[f"D{row}"] = "pos"
        ws1[f"E{row}"] = "bar"
        ws1[f"F{row}"] = "freq_bar"
        ws1[f"G{row}"] = "distance"
        ws1[f"H{row}"] = "example sentence"
        ws1[f"I{row}"] = "link"

        row += 1

        for _, de_term in tqdm(enumerate(chunk_vocab, start=1), total=len(chunk_vocab)):
            row_entity_start = row
            _id += 1

            ws1[f"A{row}"] = str(_id)
            ws1[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

            ws1[f"B{row}"] = str(de_term)
            ws1[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

            # frequency
            de_freq = de_word2freq[de_term]
            ws1[f"C{row}"].value = str(de_freq)
            ws1[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

            # most likely pos
            freq_of_most_freq_pos = max(de_pos[de_term].values())
            total_freq_pos = sum(de_pos[de_term].values())
            frac = freq_of_most_freq_pos / total_freq_pos
            most_freq_pos = [pos for pos, freq in de_pos[de_term].items() if freq == freq_of_most_freq_pos][0]
            ws1[f"D{row}"].value = f"{most_freq_pos}\n({round(frac*100, 1)} %)"
            ws1[f"D{row}"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

            # bar knn's
            knn_terms = list(de2bar_knn[de_term])
            shuffle(knn_terms)
            for bar_term, bar_term_ld in knn_terms:
                row_bar_term_start = row
                ws1[f"E{row}"].value = bar_term

                bar_term_freq = bar_word2_freq[bar_term]
                ws1[f"F{row}"].value = bar_term_freq

                ws1[f"G{row}"].value = bar_term_ld

                for rand_sent_id in bar_word2sids[bar_term]:
                    article_id = rand_sent_id.split("-")[0]
                    curid = f"https://bar.wikipedia.org/?curid={article_id}"

                    rand_sent = bar_id2sent[rand_sent_id].replace("\n", " ")
                    before, after = rand_sent.split(bar_term, 1)
                    before = before[-50:]
                    after = after[:50]
                    cell_text = CellRichText(
                        before,
                        TextBlock(InlineFont(b=True, color="1155cc"), bar_term),
                        after,
                    )

                    ws1[f"H{row}"].value = cell_text
                    ws1[f"H{row}"].alignment = Alignment(wrap_text=True) # , vertical="center")

                    ws1[f"I{row}"] = "\U0001F50E"
                    ws1[f"I{row}"].hyperlink = curid + "#:~:text=" + bar_term
                    ws1[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")

                    row += 1

                for col_idx in range(5, 8):
                    ws1.merge_cells(start_row=row_bar_term_start, end_row=row - 1, start_column=col_idx, end_column=col_idx)

                de_term_counter += 1

            for col_idx in range(1, 5):
                ws1.merge_cells(start_row=row_entity_start, end_row=row - 1, start_column=col_idx, end_column=col_idx)

        ws1.column_dimensions['a'].width = 5
        ws1.column_dimensions['b'].width = 20
        ws1.column_dimensions['c'].width = 10
        ws1.column_dimensions['d'].width = 10
        ws1.column_dimensions['e'].width = 20
        ws1.column_dimensions['h'].width = 100
        wb.save(os.path.join(output_path, f"annotations_{chunk_i}.xlsx"))
    print("Done!")


if __name__ == '__main__':
    main()
